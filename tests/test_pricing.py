"""Price ladders & the observational price/quantity slopes.

Three kinds of test, in the order they earn trust:

1. **Hand-built frames** whose ladders, posted prices, realization and slopes are
   verifiable with pencil and paper (a -1 constant-spend ladder really does fit
   to -1; a flat-price SKU really does refuse to produce a slope).
2. **Invariants on the shared 1,950-row real fixture** with lowered thresholds --
   shares bound to [0,1], the ladder reconciling to the sales frame, the posted
   price being the modal rung, the permutation p-value being a p-value.
3. **Determinism**: same input -> byte-identical CSV, and the seeded permutation
   null reproducing exactly.

The dataset-specific magnitudes (88.7% multi-price, median slopes) are NOT
asserted here -- the fixture is a sparse sample. They are measured by
`python -m retail --pricing`, quoted in the README, and drawn on plate 17.
"""

from __future__ import annotations

import numpy as np
import pandas as pd
import pytest

from retail import clean, exports, ingest, paths, pricing
from retail.__main__ import main

# Fixture-sized thresholds: the committed sample is 1,583 cleaned sales rows over
# ~1,000 SKUs, so the full-data gates (4 prices / 200 lines / 12 weeks) admit
# nothing. Lowered here on purpose, and the lowering is what is being tested.
SMALL = pricing.PricingConfig(
    min_distinct_prices=2, min_lines=3, min_weeks=2, min_week_prices=2,
    n_permutations=49, top_ladders=5,
)


def _mk(rows) -> pd.DataFrame:
    cols = ["Invoice", "StockCode", "Description", "Quantity", "InvoiceDate", "Price", "CustomerID", "Country"]
    df = pd.DataFrame(rows, columns=cols)
    df["Invoice"] = df["Invoice"].astype(str)
    df["StockCode"] = df["StockCode"].astype(str)
    df["InvoiceDate"] = pd.to_datetime(df["InvoiceDate"])
    df["CustomerID"] = df["CustomerID"].astype("Int64")
    return df


# --------------------------------------------------------------------------- #
# Hand-built cases
# --------------------------------------------------------------------------- #
def _ladder_case() -> pd.DataFrame:
    # AA sells 6 lines: four at 2.00 (the posted rung), one at 1.00, one at 4.00.
    rows = [
        ("1", "AA", "Widget AA", 10, "2011-01-04 09:00", 2.0, 1, "UK"),
        ("2", "AA", "Widget AA", 10, "2011-01-05 09:00", 2.0, 1, "UK"),
        ("3", "AA", "Widget AA", 10, "2011-01-06 09:00", 2.0, 2, "UK"),
        ("4", "AA", "Widget AA", 10, "2011-01-07 09:00", 2.0, 2, "UK"),
        ("5", "AA", "Widget AA", 20, "2011-01-11 09:00", 1.0, 3, "UK"),
        ("6", "AA", "Widget AA", 5, "2011-01-12 09:00", 4.0, 3, "UK"),
    ]
    return clean.clean(_mk(rows)).sales


def test_ladder_rungs_and_posted_price_are_hand_checkable():
    sales = _ladder_case()
    result = pricing.run_pricing_analysis(sales, SMALL)
    lad = result.ladder
    assert list(lad["Price"]) == [1.0, 2.0, 4.0]           # sorted rungs
    assert list(lad["lines"]) == [1, 4, 1]
    assert list(lad["units"]) == [20, 40, 5]
    assert list(lad["is_posted"]) == [False, True, False]  # modal BY LINES, not by units
    assert lad["unit_share"].sum() == pytest.approx(1.0)

    p = result.profile.iloc[0]
    assert p["rungs"] == 3
    assert p["posted_price"] == pytest.approx(2.0)
    # revenue 20 + 20 + 20 + 20 + 20 + 20 = 120 over 65 units
    assert p["revenue"] == pytest.approx(120.0)
    assert p["units"] == 65
    assert p["realized_price"] == pytest.approx(120.0 / 65.0)
    assert p["unit_share_below"] == pytest.approx(20 / 65)
    assert p["unit_share_at"] == pytest.approx(40 / 65)
    assert p["unit_share_above"] == pytest.approx(5 / 65)


def test_posted_price_ties_break_to_the_lower_rung():
    rows = [
        ("1", "AA", "Widget AA", 1, "2011-01-04 09:00", 3.0, 1, "UK"),
        ("2", "AA", "Widget AA", 1, "2011-01-05 09:00", 2.0, 1, "UK"),
        ("3", "AA", "Widget AA", 1, "2011-01-06 09:00", 4.0, 2, "UK"),
    ]
    sales = clean.clean(_mk(rows)).sales
    result = pricing.run_pricing_analysis(sales, SMALL)
    assert result.profile.iloc[0]["posted_price"] == pytest.approx(2.0)


def test_constant_spend_ladder_fits_exactly_minus_one():
    # Every line spends 40.00: 40/1, 20/2, 10/4, 5/8 units. No demand response
    # whatsoever -- the mechanical benchmark the module documents.
    rows = [
        ("1", "AA", "Widget AA", 40, "2011-01-04 09:00", 1.0, 1, "UK"),
        ("2", "AA", "Widget AA", 20, "2011-01-05 09:00", 2.0, 1, "UK"),
        ("3", "AA", "Widget AA", 10, "2011-01-06 09:00", 4.0, 2, "UK"),
        ("4", "AA", "Widget AA", 5, "2011-01-07 09:00", 8.0, 2, "UK"),
    ]
    sales = clean.clean(_mk(rows)).sales
    slope = pricing.within_week_slopes(sales, SMALL)
    assert slope["AA"] == pytest.approx(pricing.CONSTANT_SPEND_SLOPE)
    assert pricing.CONSTANT_SPEND_SLOPE == -1.0


def test_a_flat_price_sku_produces_no_slope_instead_of_a_fake_one():
    rows = [("1", "AA", "Widget AA", q, f"2011-01-0{d} 09:00", 2.0, 1, "UK")
            for d, q in zip(range(1, 8), [5, 9, 4, 7, 6, 8, 5], strict=True)]
    sales = clean.clean(_mk(rows)).sales
    # x has no variance -> NaN, never a fabricated 0.0
    slopes = pricing.ols_slope_by(
        sales.assign(lp=np.log(sales["Price"]), lq=np.log(sales["Quantity"])),
        "StockCode", "lp", "lq",
    )
    assert np.isnan(slopes["AA"])


def test_market_adjustment_removes_a_common_weekly_swing():
    # AA's posted price is flat-ish but weekly units swing with a market-wide
    # surge that BB shares. The adjusted series must be flatter than the raw one.
    rows = []
    for week, (price, mult) in enumerate([(2.0, 1), (2.0, 4), (1.5, 1), (1.5, 4)], start=1):
        day = f"2011-0{week}-10 09:00"
        rows.append((f"a{week}", "AA", "Widget AA", 10 * mult, day, price, 1, "UK"))
        rows.append((f"b{week}", "BB", "Widget BB", 10 * mult, day, 5.0, 2, "UK"))
    sales = clean.clean(_mk(rows)).sales
    panel = pricing.weekly_frame(sales, SMALL, pd.Index(["AA"]))
    # the market index carries the 4x swing, so the adjusted log-units barely move
    assert panel["log_units"].std() > panel["log_units_adj"].std()


def test_empty_and_thresholded_inputs_are_handled(cleaned):
    empty = pricing.run_pricing_analysis(cleaned.sales.iloc[0:0], SMALL)
    assert empty.profile.empty and empty.ladder.empty and empty.slopes.empty
    assert empty.overview["n_qualifying"] == 0
    assert pricing.fig_price_ladder(empty) is None
    assert "no SKU carried enough price variation" in pricing.headline_text(empty)
    # the full-data thresholds admit nothing from the small fixture, and that is
    # reported as an empty result rather than a crash
    strict = pricing.run_pricing_analysis(cleaned.sales)
    assert strict.overview["n_qualifying"] == 0
    assert strict.overview["n_skus"] > 0        # ...but the denominator is still real


# --------------------------------------------------------------------------- #
# Fixture-based invariants
# --------------------------------------------------------------------------- #
@pytest.fixture(scope="module")
def result(cleaned):
    return pricing.run_pricing_analysis(cleaned.sales, SMALL)


def test_fixture_has_real_price_variation_to_measure(result):
    assert result.overview["n_qualifying"] > 0
    assert result.overview["n_multi_price_skus"] > 0
    assert 0.0 <= result.overview["multi_price_share"] <= 1.0


def test_ladder_reconciles_to_the_sales_frame(cleaned, result):
    lad = result.ladder
    codes = set(lad["StockCode"])
    sub = cleaned.sales[cleaned.sales["StockCode"].isin(codes)]
    assert lad["units"].sum() == sub["Quantity"].sum()
    assert lad["revenue"].sum() == pytest.approx(sub["Revenue"].sum())
    # every rung is a price that really appears in the data, and each SKU has one posted rung
    assert set(zip(lad["StockCode"], lad["Price"], strict=True)) <= set(
        zip(sub["StockCode"], sub["Price"], strict=True)
    )
    assert (lad.groupby("StockCode", observed=True)["is_posted"].sum() == 1).all()


def test_profile_shares_are_bounded_and_sum_to_one(result):
    p = result.profile
    shares = p[["unit_share_below", "unit_share_at", "unit_share_above"]]
    assert (shares >= 0).all().all()
    assert np.allclose(shares.sum(axis=1), 1.0)
    assert (p["realized_price"] > 0).all()
    assert (p["posted_price"] > 0).all()
    assert (p["spread"] >= 1.0).all()               # p90 >= p10 by construction
    assert (p["rungs"] >= SMALL.min_distinct_prices).all()
    assert p["revenue"].is_monotonic_decreasing     # deterministic order: biggest first


def test_realized_price_is_revenue_over_units(result):
    p = result.profile
    assert np.allclose(p["realized_price"], p["revenue"] / p["units"])
    assert np.allclose(p["realization"], p["realized_price"] / p["posted_price"])


def test_realization_is_arithmetic_not_an_opportunity(result):
    # revenue_at_posted is exactly sum(posted * units) -- the same units repriced,
    # never a claim that those units would have sold at the posted price.
    o, p = result.overview, result.profile
    assert o["revenue_at_posted"] == pytest.approx((p["posted_price"] * p["units"]).sum())
    assert o["realization"] == pytest.approx(p["revenue"].sum() / o["revenue_at_posted"])


def test_slopes_are_only_fitted_where_the_thresholds_are_met(result):
    s = result.slopes
    if s.empty:
        pytest.skip("fixture too sparse for a weekly slope at these thresholds")
    assert (s["weeks"] >= SMALL.min_weeks).all()
    assert (s["week_prices"] >= SMALL.min_week_prices).all()
    assert set(s["StockCode"]) <= set(result.profile["StockCode"])


def test_permutation_p_value_is_a_p_value(result):
    s = result.slopes.dropna(subset=["perm_p"])
    if s.empty:
        pytest.skip("no fitted slopes on this fixture")
    lo = 1.0 / (SMALL.n_permutations + 1)
    assert (s["perm_p"] >= lo).all()
    assert (s["perm_p"] <= 1.0).all()


def test_permutation_null_is_reproducible(cleaned):
    a = pricing.run_pricing_analysis(cleaned.sales, SMALL).slopes
    b = pricing.run_pricing_analysis(cleaned.sales, SMALL).slopes
    pd.testing.assert_frame_equal(a, b)


def test_headline_quotes_measured_numbers_and_refuses_a_causal_claim(result):
    text = pricing.headline_text(result)
    assert f"{100 * result.overview['multi_price_share']:.1f}%" in text
    assert "elasticity" not in text.lower()
    assert "not a per-SKU price recommendation" in text or "too few weeks" in text


def test_module_never_calls_the_slope_an_elasticity_without_qualification():
    # An honesty guard with teeth: the word may appear only where it is denied.
    import pathlib

    source = pathlib.Path(pricing.__file__).read_text(encoding="utf-8")
    for line in source.splitlines():
        if "elasticity" in line.lower():
            assert any(w in line.lower() for w in ("not", "never", "proxy")), line


# --------------------------------------------------------------------------- #
# Determinism of the deliverable + the figure
# --------------------------------------------------------------------------- #
def test_csv_is_byte_identical_on_rerun(cleaned, result, tmp_path):
    p1 = pricing.write_csv(result, out_dir=tmp_path)
    b1 = p1.read_bytes()
    again = pricing.run_pricing_analysis(cleaned.sales, SMALL)
    p2 = pricing.write_csv(again, out_dir=tmp_path / "again")
    assert p2.read_bytes() == b1
    header = b1.splitlines()[0].decode("utf-8")
    assert header.startswith("StockCode,Description,rungs,lines,units,revenue,posted_price")
    assert "slope_market_adj" in header and "perm_p" in header


def test_fig_price_ladder_writes_png(result, tmp_path):
    path = pricing.fig_price_ladder(result, out_dir=tmp_path)
    assert path is not None and path.name == "price_ladder.png"
    assert path.stat().st_size > 10_000


# --------------------------------------------------------------------------- #
# CLI + exports integration
# --------------------------------------------------------------------------- #
def test_cli_pricing_skips_gracefully_without_raw_data(monkeypatch, capsys):
    def _absent(force=False):
        raise FileNotFoundError("raw data not downloaded")

    monkeypatch.setattr(ingest, "load_raw", _absent)
    assert main(["--pricing"]) == 0
    out = capsys.readouterr().out
    assert "[SKIP]" in out
    assert "download_data.py" in out


def test_cli_pricing_runs_on_fixture(monkeypatch, capsys, tmp_path, raw_fixture):
    # The fixture is far too sparse for the shipped thresholds, so the CLI must
    # still finish cleanly and say so rather than pretend it measured something.
    monkeypatch.setattr(ingest, "load_raw", lambda force=False: raw_fixture.copy())
    monkeypatch.setattr(paths, "FIGURES", tmp_path)
    monkeypatch.setattr(paths, "DELIVERABLES", tmp_path)
    assert main(["--pricing"]) == 0
    out = capsys.readouterr().out
    assert "sold at more than one price" in out
    assert "[HEADLINE]" in out
    assert (tmp_path / "price_ladder.csv").exists()


def test_excel_gains_price_ladder_sheet_only_when_present(tmp_path, cleaned):
    from openpyxl import load_workbook

    from retail import eda

    base = {
        "clean_table": cleaned.report(),
        "monthly": eda.monthly_revenue_table(cleaned.sales),
        "returns_rate": eda.returns_rate_table(cleaned.sales, cleaned.returns),
        "rfm_summary": pd.DataFrame({"Segment": ["Champions"], "Customers": [1], "RevenueShare": [1.0]}),
        "cv": pd.DataFrame({"a": [1]}),
        "cv_summary": pd.DataFrame({"model": ["x"], "mean_mase": [1.0], "worst_mase": [1.0], "fold_wins": [1]}),
        "sku_table": pd.DataFrame({"StockCode": ["1"], "Description": ["d"]}),
        "weekly": eda.monthly_revenue_table(cleaned.sales).set_index("YearMonth")["Revenue"],
    }
    xlsx_without = exports.export_excel(dict(base), tmp_path / "without.xlsx")
    wb = load_workbook(xlsx_without, read_only=True)
    assert "PriceLadder" not in wb.sheetnames
    wb.close()

    ctx = dict(base)
    ctx["pricing_result"] = pricing.run_pricing_analysis(cleaned.sales, SMALL)
    xlsx_with = exports.export_excel(ctx, tmp_path / "with.xlsx")
    wb = load_workbook(xlsx_with, read_only=True)
    assert "PriceLadder" in wb.sheetnames
    wb.close()
