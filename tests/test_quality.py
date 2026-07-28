"""Data-quality report card: exact scores, threshold bands, determinism, honesty.

Small hand-built frames pin the arithmetic; the real fixture proves the raw ->
cleaned lift on genuine Online Retail II rows. Every score here is checkable with
pencil and paper against the stated formula 100*(1 - violations/evaluated).
"""

from __future__ import annotations

import pandas as pd

from retail import exports, ingest, paths, quality
from retail.__main__ import main, retail_quality_config


# --------------------------------------------------------------------------- #
# Scoring primitives -- the bands every dimension funnels through.
# --------------------------------------------------------------------------- #
def test_score_formula_is_exact():
    assert quality._score(0, 10) == 100.0
    assert quality._score(2, 10) == 80.0
    assert quality._score(3, 200) == 98.5
    assert quality._score(8, 200) == 96.0
    assert quality._score(0, 0) == 100.0  # empty denominator -> vacuously complete


def test_status_threshold_bands():
    warn, fail = quality.WARN_BELOW, quality.FAIL_BELOW
    assert (warn, fail) == (99.5, 97.0)
    assert quality._status(100.0, warn, fail) == "pass"
    assert quality._status(99.5, warn, fail) == "pass"   # boundary is inclusive on pass
    assert quality._status(99.4, warn, fail) == "warn"
    assert quality._status(97.0, warn, fail) == "warn"   # boundary is inclusive on warn
    assert quality._status(96.9, warn, fail) == "fail"
    assert quality._status(80.0, warn, fail) == "fail"


# --------------------------------------------------------------------------- #
# Completeness -- exact score on a 10-row frame with 2 nulls.
# --------------------------------------------------------------------------- #
def test_completeness_exact_score_two_nulls_in_ten():
    df = pd.DataFrame({"k": [1, 2, None, 4, 5, None, 7, 8, 9, 10]})
    cfg = quality.QualityConfig(
        columns={"k": quality.ColumnRule(required=True)},
        plausibility_columns=(),
        dedup_columns=(),
    )
    card = quality.assess(df, cfg)
    check = next(f for f in card.findings if f.check == "non-null: k")
    assert check.score == 80.0
    assert check.status == "fail"
    assert check.count == 2 and check.total == 10
    # dimension = mean(non-null:k=80.0, overall(required)=80.0) = 80.0
    assert card.dimensions["Completeness"].score == 80.0
    assert card.dimensions["Completeness"].status == "fail"


def test_nullable_column_is_not_penalised():
    df = pd.DataFrame({"req": range(10), "opt": [1, None, None, 4, 5, 6, 7, 8, 9, 10]})
    cfg = quality.QualityConfig(
        columns={"req": quality.ColumnRule(required=True),
                 "opt": quality.ColumnRule(required=False)},
        plausibility_columns=(), dedup_columns=(),
    )
    card = quality.assess(df, cfg)
    assert card.dimensions["Completeness"].score == 100.0  # opt's nulls are allowed
    info = next(f for f in card.findings if f.check.startswith("nullable (informational): opt"))
    assert info.status == "pass" and info.count == 2


# --------------------------------------------------------------------------- #
# Per-dimension pass / warn / fail via constructed violation counts.
# --------------------------------------------------------------------------- #
def _consistency_card(neg: int, n: int = 200):
    x = [1.0] * (n - neg) + [-1.0] * neg
    df = pd.DataFrame({"x": x})
    cfg = quality.QualityConfig(
        rules=(quality.Rule("x is positive", "x", ">", 0),),
        plausibility_columns=(), dedup_columns=(),
    )
    return quality.assess(df, cfg)


def test_consistency_pass_warn_fail_bands():
    assert _consistency_card(0).dimensions["Consistency"].status == "pass"
    assert _consistency_card(3).dimensions["Consistency"].status == "warn"   # 98.5
    assert _consistency_card(8).dimensions["Consistency"].status == "fail"   # 96.0
    warn_card = _consistency_card(3)
    check = next(f for f in warn_card.findings if f.check == "x is positive")
    assert check.score == 98.5 and check.count == 3 and check.total == 200


# --------------------------------------------------------------------------- #
# Uniqueness -- known duplicate.
# --------------------------------------------------------------------------- #
def test_uniqueness_counts_exact_duplicate_rows():
    df = pd.DataFrame({"a": [1, 2, 3, 3, 5], "b": ["x", "y", "z", "z", "w"]})
    cfg = quality.QualityConfig(plausibility_columns=(), dedup_columns=("a", "b"))
    card = quality.assess(df, cfg)
    dup = next(f for f in card.findings if f.check == "duplicate rows")
    assert dup.count == 1                 # one row (3,'z') repeats
    assert dup.score == quality._score(1, 5)  # 80.0
    assert card.dimensions["Uniqueness"].status == "fail"


def test_uniqueness_key_columns():
    df = pd.DataFrame({"id": [1, 1, 2, 3], "v": [10, 20, 30, 40]})
    cfg = quality.QualityConfig(key_columns=("id",), plausibility_columns=(), dedup_columns=("id", "v"))
    card = quality.assess(df, cfg)
    key = next(f for f in card.findings if f.check.startswith("unique key: id"))
    assert key.count == 1  # id=1 repeats once


# --------------------------------------------------------------------------- #
# Validity -- declared type / range / regex / domain rules.
# --------------------------------------------------------------------------- #
def test_validity_against_declared_rules():
    df = pd.DataFrame({
        "price": [10.0, -1.0, 5.0, 0.0],          # one below min 0
        "code": ["AB12", "AB13", "xx", "AB14"],   # one fails regex
        "cat": ["a", "b", "z", "a"],              # one outside domain {a,b}
    })
    cfg = quality.QualityConfig(columns={
        "price": quality.ColumnRule(expected_type="numeric", min_value=0.0),
        "code": quality.ColumnRule(expected_type="string", regex=r"AB\d+"),
        "cat": quality.ColumnRule(expected_type="string", allowed=("a", "b")),
    }, plausibility_columns=(), dedup_columns=())
    card = quality.assess(df, cfg)
    rng = next(f for f in card.findings if f.check.startswith("range"))
    rgx = next(f for f in card.findings if f.check.startswith("regex"))
    dom = next(f for f in card.findings if f.check.startswith("domain"))
    assert rng.count == 1 and rng.score == 75.0
    assert rgx.count == 1 and rgx.score == 75.0
    assert dom.count == 1 and dom.score == 75.0
    assert card.dimensions["Validity"].status == "fail"


# --------------------------------------------------------------------------- #
# Determinism, defaults, grade cap, plausibility heuristic, empty/all-null.
# --------------------------------------------------------------------------- #
def test_determinism_same_frame_same_card():
    df = pd.DataFrame({"a": [1, 2, 2, 4], "b": [1.0, 2.0, 3.0, None]})
    cfg = retail_quality_config()
    c1, c2 = quality.assess(df, cfg), quality.assess(df, cfg)
    assert quality.render_report_card(c1) == quality.render_report_card(c2)
    assert c1.findings == c2.findings
    assert c1.overall_score == c2.overall_score and c1.grade == c2.grade


def test_default_config_path_flags_defaults_and_checks_all_columns():
    df = pd.DataFrame({"a": [1, 2, 3, 4, 5], "b": ["x", "y", None, "w", "v"]})
    card = quality.assess(df)  # no config
    assert card.used_defaults is True
    rendered = quality.render_report_card(card)
    assert "DEFAULT checks were used" in rendered
    # every column treated as required by the defaults -> b's null is a real hit
    b_check = next(f for f in card.findings if f.check == "non-null: b")
    assert b_check.count == 1
    # stated weights are the documented ones
    assert quality.DIMENSION_WEIGHTS["Completeness"] == 0.30


def test_grade_capped_by_failing_hard_dimension():
    # ~5.7% duplicate rows -> uniqueness fails -> grade capped at C even though
    # the other dimensions pass and the weighted score alone would read an A.
    base = pd.DataFrame({"a": list(range(100)), "b": list(range(100))})
    dupes = base.iloc[:6].copy()
    df = pd.concat([base, dupes], ignore_index=True)
    cfg = quality.QualityConfig(
        columns={"a": quality.ColumnRule(required=True)},  # completeness passes at 100
        plausibility_columns=(), dedup_columns=("a", "b"),
    )
    card = quality.assess(df, cfg)
    assert card.dimensions["Uniqueness"].status == "fail"
    assert card.grade == "C"                 # capped
    assert quality._grade(card.overall_score) == "A"  # score alone would be an A
    assert "capped" in quality.render_report_card(card)


def test_plausibility_is_heuristic_and_never_hard_fails():
    # a huge outlier makes plenty of "extreme" values, but plausibility must not fail
    df = pd.DataFrame({"x": [1, 2, 3, 4, 5, 6, 7, 8, 9, 10_000_000]})
    cfg = quality.QualityConfig(plausibility_columns=("x",), dedup_columns=())
    card = quality.assess(df, cfg)
    plaus = next(f for f in card.findings if f.check.startswith("outliers"))
    assert plaus.status in ("pass", "warn")   # clamped: never "fail"
    assert "HEURISTIC" in plaus.explanation
    assert card.dimensions["Plausibility"].status != "fail"


def test_empty_frame_is_graceful_and_grades_f():
    card = quality.assess(pd.DataFrame({"a": [], "b": []}))
    assert card.n_rows == 0
    assert card.grade == "F" and card.overall_score == 0.0
    rendered = quality.render_report_card(card)
    assert "0 rows" in rendered or "nothing to assess" in rendered.lower()


def test_all_null_required_column_scores_zero():
    df = pd.DataFrame({"k": [None, None, None, None]})
    cfg = quality.QualityConfig(columns={"k": quality.ColumnRule(required=True)},
                                plausibility_columns=(), dedup_columns=())
    card = quality.assess(df, cfg)
    check = next(f for f in card.findings if f.check == "non-null: k")
    assert check.score == 0.0 and check.status == "fail"


def test_not_a_certification_caveat_is_in_the_card():
    df = pd.DataFrame({"a": [1, 2, 3]})
    rendered = quality.render_report_card(quality.assess(df))
    assert quality.NOT_A_CERTIFICATION in rendered
    assert "not a certification" in rendered.lower()


# --------------------------------------------------------------------------- #
# Real fixture: the raw -> cleaned quality lift on genuine rows.
# --------------------------------------------------------------------------- #
def test_report_card_lift_on_real_fixture(raw_fixture, cleaned):
    cfg = retail_quality_config()
    before = quality.assess(raw_fixture, cfg, label="RAW")
    after = quality.assess(cleaned.sales, cfg, label="CLEANED")
    # cleaning must not make quality worse, and should lift the overall score
    assert after.overall_score >= before.overall_score
    order = quality._GRADE_ORDER
    assert order.index(after.grade) <= order.index(before.grade)
    # raw fixture genuinely has duplicates + non-positive quantities -> real findings
    assert any(f.status != "pass" for f in before.findings)
    # cleaned sales pass the hard consistency/uniqueness checks
    assert after.dimensions["Uniqueness"].status == "pass"
    assert after.dimensions["Consistency"].status == "pass"
    assert "Quality lift" in quality.compare(before, after)
    assert quality.NOT_A_CERTIFICATION in quality.render_report_card(before)


def test_findings_and_scores_frames_shapes(raw_fixture):
    card = quality.assess(raw_fixture, retail_quality_config())
    ff = quality.findings_frame(card)
    assert list(ff.columns) == ["Dimension", "Check", "Status", "Score", "Count",
                                "Evaluated", "Column", "Explanation"]
    assert len(ff) == len(card.findings)
    sf = quality.scores_frame(card)
    assert set(sf["Dimension"]) == set(quality.DIMENSIONS) | {"OVERALL"}


# --------------------------------------------------------------------------- #
# CLI: runs on the fixture; degrades gracefully when the full data is absent.
# --------------------------------------------------------------------------- #
def test_cli_report_card_runs_on_fixture(monkeypatch, capsys, tmp_path, raw_fixture):
    monkeypatch.setattr(ingest, "load_raw", lambda force=False: raw_fixture.copy())
    monkeypatch.setattr(paths, "DELIVERABLES", tmp_path)
    assert main(["--report-card"]) == 0
    out = capsys.readouterr().out
    assert "Quality lift" in out
    assert "not a certification" in out.lower()
    assert (tmp_path / "data_quality_report_card.md").exists()


def test_cli_report_card_degrades_to_fixture_without_data(monkeypatch, capsys, tmp_path):
    def _absent(force=False):
        raise FileNotFoundError("raw data not downloaded")

    monkeypatch.setattr(ingest, "load_raw", _absent)
    monkeypatch.setattr(paths, "DELIVERABLES", tmp_path)
    assert main(["--report-card"]) == 0
    out = capsys.readouterr().out
    assert "sample fixture" in out          # degraded path announced itself
    assert "Quality lift" in out


def test_quality_sheet_written_when_cards_present(tmp_path, raw_fixture, cleaned):
    from openpyxl import load_workbook

    cfg = retail_quality_config()
    ctx = {
        "quality_before": quality.assess(raw_fixture, cfg),
        "quality_after": quality.assess(cleaned.sales, cfg),
    }
    path = tmp_path / "q.xlsx"
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        exports._write_quality_sheet(writer, ctx)
    workbook = load_workbook(path, read_only=True)
    assert "DataQuality" in workbook.sheetnames
    workbook.close()
