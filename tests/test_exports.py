"""Deliverables build from fixture-sized inputs and are non-trivial files."""

from __future__ import annotations

from openpyxl import load_workbook

from retail import basket, eda, exports, forecast, ingest, lifecycle, rfm


def test_pdf_and_excel_exports_nonempty(tmp_path, raw_fixture, cleaned):
    sales, returns = cleaned.sales, cleaned.returns
    weekly = forecast.weekly_revenue(sales)
    cv = forecast.cross_validate(weekly, horizon=8, n_folds=3, step=8)
    _, rfm_summary = rfm.run_rfm(sales)
    lifecycle_result = lifecycle.lifecycle_states(sales)
    basket_result = basket.run_basket_analysis(
        sales, basket.BasketConfig(min_support=0.05, thin_support_count=3)
    )
    ctx = {
        "raw_report": ingest.raw_quality_report(raw_fixture),
        "clean_table": cleaned.report(),
        "monthly": eda.monthly_revenue_table(sales),
        "returns_rate": eda.returns_rate_table(sales, returns),
        "weekly": weekly,
        "cv": cv,
        "cv_summary": forecast.cv_summary(cv),
        "final_fold": forecast.final_fold_forecast(weekly),
        "rfm_summary": rfm_summary,
        "lifecycle_result": lifecycle_result,
        "sku_table": forecast.top_sku_forecasts(sales),
        "rules_table": basket.rules_table(basket_result.rules),
    }
    pdf = exports.export_pdf(ctx, tmp_path / "exec.pdf")
    xlsx = exports.export_excel(ctx, tmp_path / "workbook.xlsx")
    assert pdf.stat().st_size > 5_000
    assert xlsx.stat().st_size > 5_000
    workbook = load_workbook(xlsx, read_only=True)
    assert "Rules" in workbook.sheetnames
    header = [c.value for c in next(workbook["Rules"].iter_rows(max_row=1))]
    assert header[:2] == ["Antecedent", "Consequent"]
    assert "Lifecycle" in workbook.sheetnames
    lc_header = [c.value for c in next(workbook["Lifecycle"].iter_rows(min_row=3, max_row=3))]
    assert lc_header[:4] == ["month", "new", "retained", "resurrected"]
    workbook.close()
